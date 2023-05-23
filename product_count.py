# listing Each company with respect product count from a Spreadsheet

import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

product_per_supplier ={}
for product_row in range(2, product_list.max_row + 1): # product list start from 2 and end one doesn't consider, so incremented
    supplier_name = product_list.cell(product_row,4).value
    # calculation of number of products per supplier:

    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier[supplier_name]
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        print(" adding a new supplier")
        product_per_supplier[supplier_name] = 1

print(product_per_supplier)